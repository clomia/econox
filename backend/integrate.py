"""
- api와 data 모듈의 통합 로직 중 반복되는 부분을 함수로 제공합니다.
"""

import io
import asyncio
from typing import List, Dict

import numpy as np
import xarray as xr
import pandas as pd
from openpyxl.utils import get_column_letter
from fastapi import HTTPException, Request
from fastapi.responses import JSONResponse

from backend.data import fmp
from backend.data.model import Factor
from backend.data.exceptions import ElementDoesNotExist, LanguageNotSupported
from backend.calc import deinterpolate, scaling


async def lang_exception_handler(request: Request, call_next):
    """
    - LanguageNotSupported 예외를 422 HTTPException 예외로 전파시킵니다.
    - app.py에서 FastAPI 앱에 미들웨어로 등록되었습니다.
    """
    try:
        return await call_next(request)
    except LanguageNotSupported as e:
        return JSONResponse(status_code=422, content=e.message)


async def get_element(section: str, code: str):
    """Element를 가져옵니다. 존재하지 않는 경우 적절한 HTTPException을 발생시킵니다."""
    try:
        if section == "symbol":
            element = await fmp.Symbol(code).load()
        else:
            raise HTTPException(
                status_code=404,
                detail=f"element_section {section} does not exist",
            )
    except ElementDoesNotExist as e:
        raise HTTPException(status_code=404, detail=e.message)
    else:
        return element


async def get_name(
    lang: str,
    ele_sec: str,
    ele_code: str,
    fac_sec: str,
    fac_code: str,
) -> Dict[str, str]:
    """
    - feature를 유저에게 보여줄 때 필요한 구성요소의 자연어 이름 가져오기
    - return: { element: ..., factor_section: ..., factor: ... }
    """
    ele = await get_element(section=ele_sec, code=ele_code)
    factor_section = getattr(ele, ele.attr_name[fac_sec])
    factor = getattr(factor_section, fac_code)
    ele_name, factor_section_name, factor_name = await asyncio.gather(
        ele.name.en(),
        factor_section.name.trans(to=lang),
        factor.name.trans(to=lang),
    )
    return {
        "element": ele_name,
        "factor_section": factor_section_name,
        "factor": factor_name,
    }


class Feature:
    """
    - featrue 시계열을 읽고 변환하는 기능을 제공하며 예외를 적절한 HTTPException로 전파시킵니다.
    - API 함수에서 즉시 사용 가능한 시계열 데이터 불러오기 로직을 제공합니다.
    """

    def __init__(
        self,
        element_section: str,
        element_code: str,
        factor_section: str,
        factor_code: str,
    ):
        self.element_section = element_section
        self.element_code = element_code
        self.factor_section = factor_section
        self.factor_code = factor_code

    def repr_str(self):
        """피쳐를 나타내는 유니크 문자열"""
        # Python 속성명으로 사용 가능한 문자열
        return f"{self.element_section}__{self.element_code}__{self.factor_section}__{self.factor_code}"

    def repr_dict(self):
        """피쳐를 나타내는 딕셔너리 객체"""
        return {
            "element_section": self.element_section,
            "element_code": self.element_code,
            "factor_section": self.factor_section,
            "factor_code": self.factor_code,
        }

    async def to_dataset(self) -> xr.Dataset:
        """
        - 원본 형식인 Dataset 객체를 반환합니다.
        - 계산은 가급적 이 Dataset 객체를 통해 수행하세요.
        - 데이터가 존재하지 않는 경우 None 반환
        """
        element = await get_element(self.element_section, self.element_code)
        # ClientMeta 메타클래스가 만든 data_class 클래스의 인스턴스
        try:
            data_instance = getattr(element, element.attr_name[self.factor_section])
            factor: Factor = getattr(data_instance, self.factor_code)
        except KeyError:
            raise HTTPException(
                status_code=404,
                detail=f"factor_section {self.factor_section} does not exist",
            )
        except AttributeError:
            raise HTTPException(
                status_code=404, detail=f"factor_code {self.factor_code} does not exist"
            )

        return await factor.get()

    async def to_data_array(self, interpolate: bool = False) -> xr.DataArray | None:
        """데이터가 존재하지 않는 경우 None 반환"""
        if dataset := await self.to_dataset():
            return dataset.daily if interpolate else deinterpolate(dataset)

    async def to_dataframe(self, interpolate: bool = False) -> pd.DataFrame | None:
        """
        - 인덱스와 time, value 컬럼을 가진 dataframe을 반환합니다.
        - interpolate: Default[False]
            - True인 경우 모든 일(daily)가 채워진 데이터를 반환합니다.
            - False인 경우 실제 원본 데이터를 반환합니다.
        - 데이터가 존재하지 않는 경우 AssertionError
        """
        assert (data_array := await self.to_data_array(interpolate)) is not None
        data_frame = data_array.to_dataframe().reset_index()
        data_frame.t = data_frame.t.dt.date
        data_frame.columns = ["time", "value"]
        data_frame.index.name = "index"
        return data_frame

    async def to_csv(self, interpolate: bool = False) -> bytes:
        """
        - interpolate: Default[False]
            - True인 경우 모든 일(daily)가 채워진 데이터를 반환합니다.
            - False인 경우 실제 원본 데이터를 반환합니다.
        - return: 파일 데이터를 bytes로 반환합니다. 디스크를 사용하지 않습니다.
            - 데이터가 존재하지 않는 경우 AssertionError
        """
        data_frame = await self.to_dataframe(interpolate)
        data_frame.to_csv(buffer := io.BytesIO())
        return buffer.getvalue()

    async def to_xlsx(self, interpolate: bool = False) -> bytes:
        """
        - interpolate: Default[False]
            - True인 경우 모든 일(daily)가 채워진 데이터를 반환합니다.
            - False인 경우 실제 원본 데이터를 반환합니다.
        - return: 파일 데이터를 bytes로 반환합니다. 디스크를 사용하지 않습니다.
            - 데이터가 존재하지 않는 경우 AssertionError
        """
        data_frame = await self.to_dataframe(interpolate)
        sheet_name = f"econox"
        with pd.ExcelWriter(
            buffer := io.BytesIO(), date_format="YYYY-MM-DD", engine="openpyxl"
        ) as writer:
            data_frame.to_excel(writer, sheet_name=sheet_name)
            sheet = writer.sheets[sheet_name]
            # 컬럼 길이가 글자 길이보다 작으면 깨지므로 여유롭게 설정
            sheet.column_dimensions[get_column_letter(2)].width = 15
            sheet.column_dimensions[get_column_letter(3)].width = 23
        return buffer.getvalue()


class FeatureGroup:
    """
    - 피쳐들을 묶어서 그룹으로 사용 가능한 데이터를 제공합니다.
    - 시계열 데이터는 정규화되며 모든 데이터의 길이가 동일하도록 자릅니다.
    - Pandas dataframe 객체나 xlsx, csv 파일로 뽑아낼 수 있습니다.

    ```python
    group = await FeatureGroup(feature1, feature2, feature3).init()
    feature2_from_group:xr.Dataset = group[feature2] # 그룹에 맞게 슬라이싱된 Dataset
    group.dataset # 그룹 자체에 대한 Dataset
    ```
    """

    def __init__(self, *features: Feature):
        """
        - features: 그룹으로 정의할 Feature 객체들
        - 인스턴스 생성 방법:
            - `group = await FeatureGroup(...).init()`
        """
        self.src = features
        self._init = False  # init 여부

    def __getitem__(self, fe: Feature) -> xr.Dataset:
        assert self._init
        return self._dict[fe.repr_str()]  # key에 해당하는 값을 반환

    def __setitem__(self, *args):
        raise PermissionError("이 객체는 읽기 전용입니다.")

    def __delitem__(self, *args):
        raise PermissionError("이 객체는 읽기 전용입니다.")

    async def init(self):
        ds_arr = await asyncio.gather(*[fe.to_dataset() for fe in self.src])

        min_t = np.max([ds.t[0].to_numpy() for ds in ds_arr])
        max_t = np.min([ds.t[-1].to_numpy() for ds in ds_arr])

        self._dict = {  # 그룹의 개별 데이터셋 접근자 활성화
            fe.repr_str(): ds.sel(t=slice(min_t, max_t))
            for fe, ds in zip(self.src, ds_arr)
        }

        self._init = True
        return self

    async def get_columns(self, lang: str) -> List[str]:
        """
        - lang: 테이블 컬럼명에 사용할 언어
        """
        assert self._init
        names = await asyncio.gather(
            *[
                get_name(
                    lang,
                    fe.element_section,
                    fe.element_code,
                    fe.factor_section,
                    fe.factor_code,
                )
                for fe in self.src
            ]
        )
        return [
            f"[{name['element']}] {name['factor_section']}({name['factor']})"
            for name in names
        ]

    def to_dataset(self, minmax_scaling: bool = False) -> xr.Dataset:
        """
        - minmax_scaling: True인 경우 모든 값을 0에서 1사이로 Min-Max Scaling 합니다.
        """
        assert self._init
        return xr.Dataset(
            {
                fe.repr_str(): (
                    scaling(self[fe].daily) if minmax_scaling else self[fe].daily
                )
                for fe in self.src
            },
            attrs={fe.repr_str(): self[fe].attrs for fe in self.src},
        )

    async def to_dataframe(self, lang: str, minmax_scaling: bool = False):
        """
        - minmax_scaling: True인 경우 모든 값을 0에서 1사이로 Min-Max Scaling 합니다.
        - lang: 컬럼 명으로 사용할 언어
        """
        assert self._init
        data_set = self.to_dataset(minmax_scaling)
        data_frame = data_set.to_dataframe().reset_index()
        data_frame.t = data_frame.t.dt.date
        data_frame.columns = ["time"] + await self.get_columns(lang)
        data_frame.index.name = "index"
        return data_frame

    async def to_csv(self, lang: str, minmax_scaling: bool = False) -> bytes:
        """
        - minmax_scaling: True인 경우 모든 값을 0에서 1사이로 Min-Max Scaling 합니다.
        - lang: 컬럼 명으로 사용할 언어
        """
        assert self._init
        data_frame = await self.to_dataframe(lang, minmax_scaling)
        data_frame.to_csv(buffer := io.BytesIO())
        return buffer.getvalue()

    async def to_xlsx(self, lang: str, minmax_scaling: bool = False) -> bytes:
        """
        - minmax_scaling: True인 경우 모든 값을 0에서 1사이로 Min-Max Scaling 합니다.
        - lang: 컬럼 명으로 사용할 언어
        """
        assert self._init
        data_frame = await self.to_dataframe(lang, minmax_scaling)
        sheet_name = f"econox"
        with pd.ExcelWriter(
            buffer := io.BytesIO(), date_format="YYYY-MM-DD", engine="openpyxl"
        ) as writer:
            data_frame.to_excel(writer, sheet_name=sheet_name)
            sheet = writer.sheets[sheet_name]
            # 컬럼 길이를 여유롭게 설정

            for idx, name in enumerate(data_frame.columns):
                colume_letter = get_column_letter(idx + 2)
                # 23을 최소값으로 두고 이름 길이에 따라 길이 계산 (1.15는 실험적으로 찾아낸 값)
                sheet.column_dimensions[colume_letter].width = max(23, len(name) * 1.15)
            sheet.column_dimensions[get_column_letter(2)].width = 15
        return buffer.getvalue()
